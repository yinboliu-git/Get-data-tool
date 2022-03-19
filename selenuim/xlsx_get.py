#!/usr/bin/python3
#-*- codeing = utf-8 -*-
#@Time : 1/13/2021 5:20 PM
#@Author : Liu
#@File : xlsx_get.py
#@Software : PyCharm

import openpyxl
import re

# 原来，get_highest_row()和get_highest_column()在最新版的openpyxl模块中已经被删除了，取而代之的是max_row和max_column两个方法。
sp = r'.'

def get_xlsx(fielname):
    file_xlsx = openpyxl.load_workbook(fielname)
    get_data = file_xlsx[file_xlsx.sheetnames[0]]
    print(get_data)
    return get_data


def get_rows_value(xlsx_file,row):
    '''
    获取某一行的内容
    '''
    row_list = []
    for i in range(1,xlsx_file.max_column+1):
        row_list.append(xlsx_file.cell(row=row, column=i).value)
    return row_list


def get_asw(filename, nember):
    #filename = './t.xlsx'
    d  = get_xlsx(filename)
    rows = get_rows_value(d,1)
    asw = get_rows_value(d, nember+1)
    rows_sp = []
    as_dict ={}


    rows_m = rows[6:]
    asw_m = asw[6:]

    #print(asw_m)
    num = int(rows_m[len(rows_m)-1].split(sp)[0])
    print(num)
    as_list = [None]*num
    for i in range(0,num):
        as_dict[i+1] = as_list[i] = []

    print(as_list)
    for i in range(0,len(rows_m)): # 建立答案字典
        n = int(rows_m[i].split(sp)[0])-1
        rows_sp.append(n)
        asw_sp = asw_m[i]
        #print(n, int(asw_sp))
        as_list[n].append(int(asw_sp))
    print(as_dict)

    return as_dict
